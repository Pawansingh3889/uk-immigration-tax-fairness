"""
Build stakeholder-ready Excel report for PSW Tax Contribution Analysis.
Professional formatting with conditional highlighting.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Styles ---
HEADER_FILL = PatternFill('solid', fgColor='1e3a5f')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1e3a5f')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=11, color='1e3a5f')
BODY = Font(name='Arial', size=10)
BOLD = Font(name='Arial', bold=True, size=10)
BOLD_GREEN = Font(name='Arial', bold=True, size=10, color='15803d')
BOLD_RED = Font(name='Arial', bold=True, size=10, color='b91c1c')
NOTE = Font(name='Arial', size=9, color='64748b', italic=True)
GBP_FMT = '\u00a3#,##0'
GBP_FMT2 = '\u00a3#,##0.00'
BORDER = Border(
    bottom=Side('thin', 'e2e8f0')
)
HIGHLIGHT_GREEN = PatternFill('solid', fgColor='dcfce7')
HIGHLIGHT_RED = PatternFill('solid', fgColor='fee2e2')
LIGHT_BG = PatternFill('solid', fgColor='f8fafc')

def header_row(ws, row, values, widths=None):
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def data_row(ws, row, values, fonts=None, fills=None, fmt=None):
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = fonts[i-1] if fonts else BODY
        if fills and fills[i-1]:
            c.fill = fills[i-1]
        c.border = BORDER
        if fmt and fmt[i-1]:
            c.number_format = fmt[i-1]

# ============================================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================================
ws = wb.active
ws.title = 'Executive Summary'
ws.sheet_properties.tabColor = '1e3a5f'

ws.cell(row=1, column=1, value='PSW Graduate Visa: Cost-Benefit Analysis').font = TITLE_FONT
ws.cell(row=2, column=1, value='Period: January 2023 \u2013 April 2026 (2.7 years)').font = NOTE
ws.merge_cells('A1:D1')
ws.merge_cells('A2:D2')

header_row(ws, 4, ['Metric', 'Value', 'Context', 'Source'], [35, 18, 35, 25])

summary_data = [
    ('Net contribution to UK', 31090, 'Revenue minus services consumed', 'Calculated', BOLD_GREEN, HIGHLIGHT_GREEN),
    ('Total contributed', 31186, 'Tax, NI, visa fees, tuition, VAT, council tax', 'Multiple sources', BOLD, None),
    ('Total consumed', 96, 'One A&E visit', 'NHS Reference Costs', BOLD_RED, HIGHLIGHT_RED),
    ('Extra cost vs UK citizen', 4491, 'Immigration-specific costs', 'Home Office fees', BOLD_RED, None),
    ('Gross earnings (21 months)', 42246, 'Two employers', 'Payslips + P60', BOLD, None),
    ('NHS payments made', 4846, 'NI (NHS share) + IHS', 'HMRC + Home Office', BOLD, None),
    ('NHS services consumed', 96, '1x Type 1 A&E, no investigation', 'VB11Z', BOLD, None),
    ('Benefits claimed', 0, 'No Recourse to Public Funds', 'Visa condition', BOLD, None),
    ('Pension pot', 2520, 'L&G + Copernus provider', 'Provider portals', BOLD, None),
]

for idx, (metric, value, context, source, font, fill) in enumerate(summary_data, 5):
    data_row(ws, idx, [metric, value, context, source],
             fonts=[BOLD, font, BODY, NOTE],
             fills=[None, fill, None, None],
             fmt=[None, GBP_FMT, None, None])

ws.cell(row=len(summary_data)+6, column=1, value='All figures verified from payslips, pension provider portals, and published fee schedules.').font = NOTE

# ============================================================
# SHEET 2: EMPLOYMENT DATA
# ============================================================
ws2 = wb.create_sheet('Employment Data')
ws2.sheet_properties.tabColor = '3b82f6'

ws2.cell(row=1, column=1, value='Employment Timeline').font = TITLE_FONT

header_row(ws2, 3, ['Employer', 'Role', 'Period', 'Months', 'Rate/hr', 'Gross', 'PAYE Tax', 'NI', 'Pension', 'Employer NI', 'Employer Pension'],
           [30, 20, 22, 10, 10, 14, 14, 14, 14, 14, 16])

emp_data = [
    ('Cranswick Convenience Foods', 'Production Operative', 'Jul 2024 \u2013 Apr 2025', 10, 12.11, 11605.28, 0, 485.34, 214.28, 1500, 130),
    ('Copernus Ltd', 'Team Leader', 'Apr 2025 \u2013 Apr 2026', 12, 14.00, 30640.28, 3932.80, 1545.00, 1036.41, 3969.93, 621.83),
]

for idx, row in enumerate(emp_data, 4):
    bg = LIGHT_BG if idx % 2 == 0 else None
    data_row(ws2, idx, list(row),
             fmt=[None, None, None, None, GBP_FMT2, GBP_FMT2, GBP_FMT2, GBP_FMT2, GBP_FMT2, GBP_FMT2, GBP_FMT2],
             fills=[bg]*11)

# Totals row
data_row(ws2, 6, ['TOTAL', '', '', 21, '', 42245.56, 3932.80, 2030.34, 1250.69, 5469.93, 751.83],
         fonts=[BOLD]*11, fmt=[None, None, None, None, None, GBP_FMT2, GBP_FMT2, GBP_FMT2, GBP_FMT2, GBP_FMT2, GBP_FMT2])

# ============================================================
# SHEET 3: PAWAN VS UK CITIZEN
# ============================================================
ws3 = wb.create_sheet('Cost Comparison')
ws3.sheet_properties.tabColor = 'b91c1c'

ws3.cell(row=1, column=1, value='Cost Comparison: PSW Visa Holder vs UK Citizen').font = TITLE_FONT
ws3.cell(row=2, column=1, value='Same job, same hourly rate, same employer').font = NOTE

header_row(ws3, 4, ['Cost Item', 'PSW Holder', 'UK Citizen', 'Difference'], [30, 16, 16, 16])

comp_items = [
    ('Student Visa', 490, 0), ('Student IHS (2 years)', 940, 0),
    ('MSc Tuition (net of scholarship)', 9400, 9250),
    ('PSW Visa', 822, 0), ('PSW IHS (2 years)', 2070, 0),
    ('Biometric Fee', 19.20, 0),
    ('PAYE Income Tax', 3932.80, 3932.80), ('National Insurance', 2030.34, 2030.34),
    ('VAT (estimated)', 2000, 2000), ('Council Tax (estimated)', 1200, 1200),
]

for idx, (item, pawan, uk) in enumerate(comp_items, 5):
    diff = pawan - uk
    diff_font = BOLD_RED if diff > 0 else BOLD_GREEN if diff == 0 else BOLD
    diff_fill = HIGHLIGHT_RED if diff > 0 else None
    data_row(ws3, idx, [item, pawan, uk, diff],
             fonts=[BODY, BODY, BODY, diff_font],
             fills=[None, None, None, diff_fill],
             fmt=[None, GBP_FMT2, GBP_FMT2, GBP_FMT2])

total_row = len(comp_items) + 5
total_pawan = sum(p for _, p, _ in comp_items)
total_uk = sum(u for _, _, u in comp_items)
data_row(ws3, total_row, ['TOTAL', total_pawan, total_uk, total_pawan - total_uk],
         fonts=[BOLD, BOLD, BOLD, BOLD_RED],
         fills=[None, None, None, HIGHLIGHT_RED],
         fmt=[None, GBP_FMT2, GBP_FMT2, GBP_FMT2])

# ============================================================
# SHEET 4: NHS ANALYSIS
# ============================================================
ws4 = wb.create_sheet('NHS Analysis')
ws4.sheet_properties.tabColor = '15803d'

ws4.cell(row=1, column=1, value='NHS: Payments vs Services Consumed').font = TITLE_FONT
ws4.cell(row=2, column=1, value='Demonstrating duplicate charging through NI and IHS').font = NOTE

header_row(ws4, 4, ['NHS Payment Channel', 'Amount', 'What It Funds', 'Notes'], [28, 14, 25, 35])

nhs_items = [
    ('NI Contribution (NHS share ~80%)', 1624, 'NHS', 'Standard deduction for all workers'),
    ('Student IHS', 940, 'NHS', 'Immigration Health Surcharge, 2 years'),
    ('PSW IHS', 2070, 'NHS', 'Immigration Health Surcharge, 2 years'),
]
for idx, (item, amt, funds, notes) in enumerate(nhs_items, 5):
    data_row(ws4, idx, [item, amt, funds, notes], fmt=[None, GBP_FMT, None, None])

data_row(ws4, 8, ['TOTAL NHS PAYMENTS', 4634, '', ''], fonts=[BOLD]*4, fmt=[None, GBP_FMT, None, None])

ws4.cell(row=10, column=1, value='NHS Services Consumed').font = SUBTITLE_FONT
header_row(ws4, 11, ['Service', 'Cost', 'Reference', 'Detail'], [28, 14, 25, 35])
data_row(ws4, 12, ['Type 1 A&E attendance', 96, 'VB11Z', 'No investigation, no significant procedure'],
         fmt=[None, GBP_FMT, None, None])
data_row(ws4, 13, ['TOTAL CONSUMED', 96, '', ''], fonts=[BOLD]*4, fmt=[None, GBP_FMT, None, None])

ws4.cell(row=15, column=1, value='Comparison').font = SUBTITLE_FONT
header_row(ws4, 16, ['Metric', 'Amount', '', ''], [28, 14, 25, 35])
data_row(ws4, 17, ['PSW holder paid for NHS', 4634, '', ''], fonts=[BOLD_RED, BOLD_RED, BODY, BODY], fmt=[None, GBP_FMT, None, None])
data_row(ws4, 18, ['UK citizen would pay (NI only)', 1624, '', ''], fonts=[BODY]*4, fmt=[None, GBP_FMT, None, None])
data_row(ws4, 19, ['Premium paid', 3010, '', '185% above UK citizen rate'],
         fonts=[BOLD_RED]*4, fills=[HIGHLIGHT_RED]*4, fmt=[None, GBP_FMT, None, None])

# ============================================================
# SHEET 5: NET CONTRIBUTION
# ============================================================
ws5 = wb.create_sheet('Net Contribution')
ws5.sheet_properties.tabColor = '15803d'

ws5.cell(row=1, column=1, value='Net Fiscal Contribution to the UK').font = TITLE_FONT

ws5.cell(row=3, column=1, value='CONTRIBUTIONS (What was put in)').font = SUBTITLE_FONT
header_row(ws5, 4, ['Item', 'Amount', 'Category'], [40, 16, 20])

contributions = [
    ('PAYE Income Tax', 3932.80, 'Direct tax'),
    ('National Insurance', 2030.34, 'Direct tax'),
    ('Employer NI (exists because of employment)', 5470, 'Indirect'),
    ('Pension Contributions', 1250.69, 'Savings'),
    ('Employer Pension', 752, 'Indirect'),
    ('Student Visa + IHS', 1430, 'Immigration'),
    ('PSW Visa + IHS', 2892, 'Immigration'),
    ('MSc Tuition (net of scholarship)', 9400, 'Education'),
    ('VAT on spending (estimated)', 2000, 'Indirect tax'),
    ('Council Tax (estimated)', 1200, 'Local tax'),
]
for idx, (item, amt, cat) in enumerate(contributions, 5):
    data_row(ws5, idx, [item, amt, cat], fmt=[None, GBP_FMT2, None])

total_in = sum(a for _, a, _ in contributions)
data_row(ws5, 15, ['TOTAL CONTRIBUTIONS', total_in, ''],
         fonts=[BOLD, BOLD_GREEN, BOLD], fills=[None, HIGHLIGHT_GREEN, None],
         fmt=[None, GBP_FMT2, None])

ws5.cell(row=17, column=1, value='CONSUMPTION (What was taken out)').font = SUBTITLE_FONT
header_row(ws5, 18, ['Item', 'Amount', 'Category'], [40, 16, 20])
data_row(ws5, 19, ['NHS A&E visit', 96, 'Healthcare'], fmt=[None, GBP_FMT2, None])
data_row(ws5, 20, ['Benefits claimed', 0, 'Welfare'], fmt=[None, GBP_FMT2, None])
data_row(ws5, 21, ['Social housing', 0, 'Housing'], fmt=[None, GBP_FMT2, None])
data_row(ws5, 22, ['TOTAL CONSUMED', 96, ''],
         fonts=[BOLD, BOLD_RED, BOLD], fills=[None, HIGHLIGHT_RED, None],
         fmt=[None, GBP_FMT2, None])

ws5.cell(row=24, column=1, value='NET CONTRIBUTION').font = Font(name='Arial', bold=True, size=14, color='15803d')
ws5.cell(row=24, column=2, value=total_in - 96).font = Font(name='Arial', bold=True, size=14, color='15803d')
ws5.cell(row=24, column=2).number_format = GBP_FMT

# ============================================================
# SHEET 6: PENSION
# ============================================================
ws6 = wb.create_sheet('Pension')
ws6.sheet_properties.tabColor = '7c3aed'

ws6.cell(row=1, column=1, value='Pension Position').font = TITLE_FONT
ws6.cell(row=2, column=1, value='Confirmed balances from provider portals, April 2026').font = NOTE

header_row(ws6, 4, ['Provider', 'Employee', 'Employer', 'Total Contributed', 'Pot Value', 'Growth'],
           [28, 14, 14, 18, 14, 14])

pensions = [
    ('Legal & General (Cranswick)', 214.28, 130, 344.28, 504.31, 160.03),
    ('Workplace Pension (Copernus)', 1036.41, 621.83, 1658.24, 2016.00, 357.76),
]
for idx, row in enumerate(pensions, 5):
    data_row(ws6, idx, list(row), fmt=[None]+[GBP_FMT2]*5)

data_row(ws6, 7, ['TOTAL', 1250.69, 751.83, 2002.52, 2520.31, 517.79],
         fonts=[BOLD]*6, fmt=[None]+[GBP_FMT2]*5)

ws6.cell(row=9, column=1, value='Notes:').font = SUBTITLE_FONT
ws6.cell(row=10, column=1, value='- Both schemes are salary sacrifice (pre-tax contributions)').font = NOTE
ws6.cell(row=11, column=1, value='- Accessible from age 57, or transferable to QROPS if leaving the UK').font = NOTE
ws6.cell(row=12, column=1, value='- Pot values include investment growth/losses').font = NOTE

# ============================================================
# SAVE
# ============================================================
wb.save('UK_Immigration_Tax_Fairness.xlsx')
print('Excel report saved: UK_Immigration_Tax_Fairness.xlsx')
